import os
import threading
import uuid
import re
from io import BytesIO
from flask import Flask, request, render_template, redirect, url_for, send_file
from datetime import datetime
import requests
import openpyxl

app = Flask(__name__)

# Folder na wyniki - twórz, jeśli nie istnieje
wyniki_folder = "wyniki"
os.makedirs(wyniki_folder, exist_ok=True)

# --- Funkcja do sprawdzania NIP w rejestrze VAT ---
def sprawdz_nip_w_vat(nip):
    nip = re.sub(r"\D", "", str(nip))  # usuń wszystko poza cyframi
    if not nip.isdigit() or len(nip) != 10:
        return nip, "Nieprawidłowy NIP", "Błąd"

    base_url = "https://wl-api.mf.gov.pl/api/search/nip/"
    today = datetime.today().strftime('%Y-%m-%d')
    url = f"{base_url}{nip}?date={today}"
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'RequestId': str(uuid.uuid4())
    }

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            result = response.json()
            subject = result.get('result', {}).get('subject')
            if subject:
                nazwa = subject.get('name', 'Brak danych')
                status_vat = subject.get('statusVat', 'Nieznany')
                return nip, nazwa, status_vat
            else:
                return nip, "Nie znaleziono w rejestrze", "Brak"
        else:
            return nip, "Błąd odpowiedzi", f"Kod {response.status_code}"
    except Exception as e:
        return nip, "Błąd zapytania", str(e)

# --- Wczytanie NIP-ów z pliku Excel ---
def wczytaj_nipy_z_excel(plik):
    wb = openpyxl.load_workbook(plik)
    ws = wb.active
    nipy = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nip = row[0]
        if nip:
            nipy.append(str(nip).strip())
    return nipy

# --- Generowanie pliku Excel z wynikami ---
def generuj_excel(wyniki):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["NIP", "Nazwa podmiotu", "Status VAT"])
    for w in wyniki:
        ws.append(w)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

# --- Słownik do przechowywania statusów zadań ---
statusy = {}

# --- Funkcja przetwarzająca plik w tle ---
def przetworz_plik(task_id, plik):
    try:
        nipy = wczytaj_nipy_z_excel(plik)
        wyniki = [sprawdz_nip_w_vat(nip) for nip in nipy]
        excel = generuj_excel(wyniki)
        sciezka = os.path.join(wyniki_folder, f"{task_id}.xlsx")
        with open(sciezka, "wb") as f:
            f.write(excel.read())
        statusy[task_id] = ("gotowe", sciezka)
    except Exception as e:
        statusy[task_id] = ("blad", str(e))

# --- Główna strona ---
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        nip_input = request.form.get("nip", "").strip()
        plik = request.files.get("plik")

        if nip_input:
            wynik = sprawdz_nip_w_vat(nip_input)
            return render_template("wyniki.html", wyniki=[wynik])

        elif plik and plik.filename.endswith(".xlsx"):
            task_id = str(uuid.uuid4())
            statusy[task_id] = ("w trakcie", None)
            threading.Thread(target=przetworz_plik, args=(task_id, plik)).start()
            return redirect(url_for("status", task_id=task_id))

        else:
            return render_template("index.html", blad="Wpisz NIP lub załaduj plik .xlsx.")
    return render_template("index.html")

# --- Status przetwarzania ---
@app.route("/status/<task_id>")
def status(task_id):
    stan = statusy.get(task_id, ("nieznany", None))
    if stan[0] == "w trakcie":
        return f"⏳ Trwa przetwarzanie... <br><a href='{url_for('status', task_id=task_id)}'>Odśwież</a>"
    elif stan[0] == "gotowe":
        return f"✅ Gotowe! <a href='{url_for('pobierz_wynik', task_id=task_id)}'>Pobierz wyniki</a>"
    elif stan[0] == "blad":
        return f"❌ Błąd: {stan[1]}"
    else:
        return "Nie znaleziono zadania."

# --- Pobieranie gotowego pliku ---
@app.route("/pobierz/<task_id>")
def pobierz_wynik(task_id):
    stan = statusy.get(task_id, (None, None))
    if stan[0] == "gotowe":
        return send_file(stan[1], as_attachment=True)
    return "Plik niedostępny."

if __name__ == "__main__":
    app.run(debug=True)
